import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path.cwd()
WORKBOOK = next(Path(r"C:\Users\10524\Desktop\codex_realasset\Project\03_Logi_Leasing_Dashboard").glob("*260414*.xlsx"))
OUT_DIR = ROOT / "qa-artifacts" / "logistics-gate6" / "xlsx-cell-coverage-check-20260514"
OUT_DIR.mkdir(parents=True, exist_ok=True)

EXPECTED = {
    "DB_일반": {"cell_count": 6216, "non_empty_count": 4909, "formula_count": 763, "max_row": 74, "max_col": 84},
    "DB_히스토리 누적": {"cell_count": 3382, "non_empty_count": 2885, "formula_count": 172, "max_row": 178, "max_col": 19},
    "Log": {"cell_count": 518, "non_empty_count": 321, "formula_count": 0, "max_row": 37, "max_col": 14},
    "Meta_데이터 항목 설명": {"cell_count": 3476, "non_empty_count": 385, "formula_count": 6, "max_row": 79, "max_col": 44},
    "자산_담당자 연결": {"cell_count": 160, "non_empty_count": 127, "formula_count": 0, "max_row": 20, "max_col": 8},
}

wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
rows = []
for sheet_name, expected in EXPECTED.items():
    ws = wb[sheet_name]
    non_empty = 0
    formulas = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = cell.value
            if value not in (None, ""):
                non_empty += 1
            if isinstance(value, str) and value.startswith("="):
                formulas += 1
    cell_count = ws.max_row * ws.max_column
    status = "pass" if (
        cell_count == expected["cell_count"]
        and non_empty == expected["non_empty_count"]
        and formulas == expected["formula_count"]
    ) else "fail"
    rows.append({
        "sheet_name": sheet_name,
        "xlsx_cell_count": cell_count,
        "supabase_cell_count": expected["cell_count"],
        "xlsx_non_empty_count": non_empty,
        "supabase_non_empty_count": expected["non_empty_count"],
        "xlsx_formula_count": formulas,
        "supabase_formula_count": expected["formula_count"],
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "status": status,
    })

result = {
    "generated_at": datetime.now(timezone.utc).isoformat(),
    "workbook": str(WORKBOOK),
    "mutation_performed": False,
    "sheet_count_checked": len(rows),
    "all_pass": all(row["status"] == "pass" for row in rows),
    "rows": rows,
}

(OUT_DIR / "result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
summary = [
    "# XLSX Cell Coverage Check - 2026-05-14",
    "",
    f"- workbook: {WORKBOOK}",
    "- mutation_performed: false",
    f"- sheet_count_checked: {len(rows)}",
    f"- all_pass: {str(result['all_pass']).lower()}",
    "",
    "| sheet | xlsx cells | Supabase cells | xlsx non-empty | Supabase non-empty | xlsx formulas | Supabase formulas | status |",
    "|---|---:|---:|---:|---:|---:|---:|---|",
]
for row in rows:
    summary.append(
        f"| {row['sheet_name']} | {row['xlsx_cell_count']} | {row['supabase_cell_count']} | "
        f"{row['xlsx_non_empty_count']} | {row['supabase_non_empty_count']} | "
        f"{row['xlsx_formula_count']} | {row['supabase_formula_count']} | {row['status']} |"
    )
(OUT_DIR / "summary.md").write_text("\n".join(summary) + "\n", encoding="utf-8")
print(json.dumps(result, ensure_ascii=False, indent=2))
